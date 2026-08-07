"""Export formats — step 10 of the workflow."""

from __future__ import annotations

import io
import json
import zipfile
from datetime import UTC, datetime

import pytest

from app.core.errors import ExportFormatError
from app.core.text import format_timestamp, slugify_filename
from app.services.export import (
    ExportDocument,
    ExportSegment,
    available_formats,
    export_many,
    export_one,
    format_names,
)


def _document(title: str = "How to write a hook", *, segments: bool = True) -> ExportDocument:
    return ExportDocument(
        transcript_id="t1",
        video_id="v1",
        title=title,
        platform="youtube",
        source_url="https://www.youtube.com/watch?v=abc",
        author="Creator Name",
        text="First line of the transcript. Second line of the transcript.",
        segments=[
            ExportSegment(index=0, start=0.0, end=3.5, text="First line of the transcript."),
            ExportSegment(index=1, start=3.5, end=7.25, text="Second line of the transcript."),
        ]
        if segments
        else [],
        duration_seconds=7.25,
        language="en",
        provider="stub",
        model="stub-v1",
        word_count=10,
        created_at=datetime(2026, 1, 1, tzinfo=UTC),
    )


def test_all_seven_formats_are_available() -> None:
    assert set(format_names()) == {"txt", "docx", "md", "xlsx", "json", "srt", "vtt"}
    assert len(available_formats()) == 7


@pytest.mark.parametrize("format_name", ["txt", "docx", "md", "xlsx", "json", "srt", "vtt"])
def test_every_format_renders_non_empty_bytes(format_name: str) -> None:
    content, filename, content_type = export_one(_document(), format_name)
    assert content
    assert filename.endswith(f".{format_name}")
    assert content_type


def test_txt_contains_metadata_and_body() -> None:
    content, _, _ = export_one(_document(), "txt")
    text = content.decode()
    assert "How to write a hook" in text
    assert "Creator Name" in text
    assert "First line of the transcript." in text


def test_markdown_has_heading_and_source_link() -> None:
    text = export_one(_document(), "md")[0].decode()
    assert text.startswith("# How to write a hook")
    assert "https://www.youtube.com/watch?v=abc" in text


def test_json_round_trips_with_segments() -> None:
    payload = json.loads(export_one(_document(), "json")[0])
    assert payload["title"] == "How to write a hook"
    assert payload["word_count"] == 10
    assert len(payload["segments"]) == 2
    assert payload["segments"][0]["start"] == 0.0


def test_srt_is_well_formed() -> None:
    text = export_one(_document(), "srt")[0].decode()
    assert text.startswith("1\n")
    assert "00:00:00,000 --> 00:00:03,500" in text
    assert "00:00:03,500 --> 00:00:07,250" in text


def test_vtt_has_the_required_header_and_dot_separator() -> None:
    text = export_one(_document(), "vtt")[0].decode()
    assert text.startswith("WEBVTT")
    assert "00:00:00.000 --> 00:00:03.500" in text


def test_docx_is_a_real_openxml_package() -> None:
    content = export_one(_document(), "docx")[0]
    assert content[:2] == b"PK"
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        assert "word/document.xml" in archive.namelist()


def test_xlsx_has_both_sheets() -> None:
    from openpyxl import load_workbook

    content = export_one(_document(), "xlsx")[0]
    workbook = load_workbook(io.BytesIO(content))
    assert workbook.sheetnames == ["Transcripts", "Segments"]
    assert workbook["Transcripts"].cell(row=2, column=1).value == "How to write a hook"
    assert workbook["Segments"].max_row == 3  # header + two segments


def test_timed_formats_refuse_a_transcript_without_segments() -> None:
    document = _document(segments=False)
    for format_name in ("srt", "vtt"):
        with pytest.raises(ExportFormatError):
            export_one(document, format_name)


def test_unknown_format_is_rejected_with_the_supported_list() -> None:
    with pytest.raises(ExportFormatError) as exc:
        export_one(_document(), "pdf")
    assert "txt" in exc.value.message


def test_bulk_text_format_produces_one_readable_file() -> None:
    """A batch is usually one piece of research, so it defaults to one file."""
    documents = [_document("First video"), _document("Second video")]
    content, filename, content_type = export_many(documents, "txt")
    assert filename.endswith(".txt")
    assert content_type.startswith("text/plain")

    text = content.decode("utf-8")
    assert "CONTENTS" in text
    # Each title appears twice: once in the contents list, once as its section.
    for title in ("First video", "Second video"):
        assert text.count(title) == 2


def test_bulk_docx_is_one_document() -> None:
    content, filename, _ = export_many([_document("A"), _document("B")], "docx")
    assert filename.endswith(".docx")
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        body = archive.read("word/document.xml").decode("utf-8")
    assert "Combined transcripts" in body
    assert "A" in body and "B" in body


def test_bulk_text_format_can_still_ask_for_separate_files() -> None:
    documents = [_document("First video"), _document("Second video")]
    content, filename, content_type = export_many(documents, "txt", combine=False)
    assert content_type == "application/zip"
    assert filename.endswith(".zip")
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        assert sorted(archive.namelist()) == ["First_video.txt", "Second_video.txt"]


def test_bulk_zip_deduplicates_identical_titles() -> None:
    """Two videos with the same title must not overwrite each other."""
    content, _, _ = export_many([_document("Same"), _document("Same")], "md", combine=False)
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        assert len(archive.namelist()) == 2


def test_subtitle_formats_stay_separate_even_when_combining() -> None:
    """SRT timings all start at zero, so a concatenation would not play."""
    content, filename, content_type = export_many(
        [_document("A"), _document("B")], "srt", combine=True
    )
    assert content_type == "application/zip"
    assert filename.endswith(".zip")


def test_bulk_xlsx_is_one_workbook_not_a_zip_of_workbooks() -> None:
    from openpyxl import load_workbook

    content, filename, _ = export_many([_document("A"), _document("B")], "xlsx")
    assert filename.endswith(".xlsx")
    workbook = load_workbook(io.BytesIO(content))
    assert workbook["Transcripts"].max_row == 3  # header + two rows


def test_bulk_json_is_one_document() -> None:
    content, filename, _ = export_many([_document("A"), _document("B")], "json")
    payload = json.loads(content)
    assert filename.endswith(".json")
    assert payload["count"] == 2


def test_bulk_timed_format_skips_transcripts_without_segments() -> None:
    documents = [_document("Has segments"), _document("No segments", segments=False)]
    content, _, _ = export_many(documents, "srt")
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        assert archive.namelist() == ["Has_segments.srt"]


def test_bulk_timed_format_errors_when_nothing_qualifies() -> None:
    with pytest.raises(ExportFormatError):
        export_many([_document("No segments", segments=False)], "srt")


@pytest.mark.parametrize(
    ("title", "expected"),
    [
        ("../../etc/passwd", "etcpasswd"),
        ("Hook / CTA \\ Offer", "Hook_CTA_Offer"),
        ("Reel: 5 tips!", "Reel_5_tips"),
        ("日本語タイトル", "transcript"),
        ("", "transcript"),
    ],
)
def test_filenames_are_sanitised(title: str, expected: str) -> None:
    """User-controlled titles reach ZIP entries and headers — they must be safe."""
    assert slugify_filename(title) == expected


def test_timestamp_formatting_handles_hours() -> None:
    assert format_timestamp(3723.456, separator=",") == "01:02:03,456"
    assert format_timestamp(0.0, separator=".") == "00:00:00.000"
